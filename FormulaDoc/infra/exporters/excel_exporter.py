from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from modules.img_to_excel.schemas import WorkbookData


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")


def export_workbook_to_excel(workbook: WorkbookData, output_path: Path) -> None:
    if not workbook.sheets:
        raise ValueError("没有可导出的 sheet。")

    wb = Workbook()
    first_sheet = True

    for sheet_data in workbook.sheets:
        if first_sheet:
            ws = wb.active
            first_sheet = False
        else:
            ws = wb.create_sheet()

        ws.title = sanitize_sheet_title(sheet_data.sheet_name)
        row_index = 1

        if sheet_data.headers:
            for column_index, value in enumerate(sheet_data.headers, start=1):
                cell = ws.cell(row=row_index, column=column_index, value=value)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = HEADER_FILL
            row_index += 1

        for row in sheet_data.rows:
            for column_index, value in enumerate(row, start=1):
                cell = ws.cell(row=row_index, column=column_index, value=value)
                cell.alignment = Alignment(vertical="center")
            row_index += 1

        apply_column_widths(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def sanitize_sheet_title(raw_title: str) -> str:
    invalid_chars = set("\\/*?:[]")
    cleaned = "".join(char for char in raw_title if char not in invalid_chars).strip()
    if not cleaned:
        cleaned = "Sheet1"
    return cleaned[:31]


def apply_column_widths(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 40)
